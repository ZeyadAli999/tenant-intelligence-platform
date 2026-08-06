"""Read-only XLSX parser with formula values never executed."""

from pathlib import Path

from openpyxl import load_workbook

from app.config import Settings
from services.documents.parsers.base import ParsedDocument, ParsedElement
from services.documents.upload_security import FileValidationError


class XLSXParser:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    def parse(self, path: Path) -> ParsedDocument:
        try:
            with path.open("rb") as source:
                workbook = load_workbook(
                    source, read_only=True, data_only=True, keep_links=False
                )
                elements: list[ParsedElement] = []
                cells = rows = 0
                for sheet in workbook.worksheets:
                    batch: list[str] = []
                    start = 1
                    for row_number, row in enumerate(
                        sheet.iter_rows(values_only=True), 1
                    ):
                        rows += 1
                        cells += len(row)
                        if (
                            rows > self.settings.document_max_spreadsheet_rows
                            or cells > self.settings.document_max_spreadsheet_cells
                        ):
                            raise FileValidationError("CONTENT_LIMIT_EXCEEDED")
                        batch.append(
                            " | ".join(
                                "" if value is None else str(value) for value in row
                            )
                        )
                        if len(batch) == 50:
                            elements.append(
                                ParsedElement(
                                    "\n".join(batch),
                                    sheet_name=sheet.title,
                                    row_start=start,
                                    row_end=row_number,
                                )
                            )
                            batch, start = [], row_number + 1
                    if batch:
                        elements.append(
                            ParsedElement(
                                "\n".join(batch),
                                sheet_name=sheet.title,
                                row_start=start,
                                row_end=start + len(batch) - 1,
                            )
                        )
                workbook.close()
            return ParsedDocument(tuple(elements))
        except FileValidationError:
            raise
        except Exception as exc:
            raise FileValidationError("PARSING_FAILED") from exc
